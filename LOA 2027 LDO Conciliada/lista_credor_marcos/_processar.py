"""
Processa lista de precatorios LOA 2027 enviada por Marcos.
Aplica regra art. 100 §5º CF + cruza com LOA 2026 + gera XLSX consolidado.
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

DEST = r"c:\Users\MarcosCosta\OneDrive - CTS Brasil\Área de Trabalho\ClaudeCode\AuraLOA\LOA 2027 LDO Conciliada\lista_credor_marcos"
LOA2026 = r"c:\Users\MarcosCosta\OneDrive - CTS Brasil\Área de Trabalho\ClaudeCode\ArquivosLOA\Arquivos Precatorio\LOA_2026\precatorios_extraidos.csv"
LOA2026_CONC = r"c:\Users\MarcosCosta\OneDrive - CTS Brasil\Área de Trabalho\ClaudeCode\ArquivosLOA\Arquivos Precatorio\LOA_2026\precatorios_ACIMA_10M_SIOP_conciliado_v2.csv"

# 1) Dados brutos
raw = [
    (None,   "1687386020254019198", "119.179.000,05"),
    ("1705", "1683168520254019198", "25.679.565,73"),
    ("1695", "1681323220254019198", "126.740.315,61"),
    ("1574", "1666929820254019198", "23.594.343,72"),
    ("1659", "1678153420254019198", "42.296.067,54"),
    ("1660", "1678327020254019198", "21.483.826,92"),
    ("1608", "1670896020254019198", "140.053.568,72"),
    ("1609", "1671918220254019198", "68.399.404,30"),
    ("1589", "1668497120254019198", "99.028.489,41"),
    ("1590", "1668505620254019198", "135.848.606,40"),
    ("1591", "1668514120254019198", "110.795.906,84"),
    ("1592", "1668522620254019198", "120.332.955,10"),
    ("4",    "2152261020244019198", "205.286,80"),
    ("5",    "2152373920244019198", "187.518,56"),
    ("6",    "2154383120244019198", "121.056,71"),
    ("7",    "2154391620244019198", "151.362,67"),
    ("8",    "2154409820244019198", "371.922,65"),
    ("9",    "2154418320244019198", "293.523,38"),
    ("10",   "2154426820244019198", "40.969,52"),
]
df = pd.DataFrame(raw, columns=["ordem_oficio", "cnj_19digitos", "valor_brl_str"])

# 2) Formatar CNJ canonico
def format_cnj(d19: str) -> str:
    """19 digitos -> NNNNNNN-DD.AAAA.J.TR.OOOO (adiciona leading zero)"""
    d20 = d19.zfill(20)
    return f"{d20[0:7]}-{d20[7:9]}.{d20[9:13]}.{d20[13:14]}.{d20[14:16]}.{d20[16:20]}"

def cnj_ano(d19: str) -> int:
    return int(d19.zfill(20)[9:13])

def cnj_segmento(d19: str) -> str:
    seg = d19.zfill(20)[13:14]
    return {"4": "Justiça Federal", "5": "Justiça do Trabalho", "8": "Justiça Estadual"}.get(seg, seg)

def cnj_tribunal(d19: str) -> str:
    seg = d19.zfill(20)[13:14]
    trib = d19.zfill(20)[14:16]
    if seg == "4":
        return f"TRF{int(trib)}"
    return f"seg{seg}/trib{trib}"

def cnj_origem(d19: str) -> str:
    return d19.zfill(20)[16:20]

df["cnj_canonico"] = df["cnj_19digitos"].apply(format_cnj)
df["ano_autuacao"] = df["cnj_19digitos"].apply(cnj_ano)
df["segmento"] = df["cnj_19digitos"].apply(cnj_segmento)
df["tribunal"] = df["cnj_19digitos"].apply(cnj_tribunal)
df["origem"] = df["cnj_19digitos"].apply(cnj_origem)

# 3) Valor numerico
def br_to_float(s: str) -> float:
    return float(s.replace(".", "").replace(",", "."))

df["valor_brl"] = df["valor_brl_str"].apply(br_to_float)

# 4) Aplicar regra art. 100 §5º CF
# Regra: Precatorio autuado ate 02/04/N -> paga em N+1
# Logo: autuacao ate 02/04/2025 -> LOA 2026 ; autuacao ate 02/04/2026 -> LOA 2027
# CNJ ano = ano de AUTUACAO no TRF
def loa_provavel(ano_autuacao: int) -> str:
    """LOA prevista pela regra art. 100 §5º.

    Como nao temos a data exata de autuacao (so o ano do CNJ), o ano N do CNJ
    indica que o precatorio foi autuado em N. Pela regra:
      - autuacao ate 02/04/N -> paga em N+1
      - autuacao apos 02/04/N -> paga em N+2

    Logo, o ano do CNJ=N pode cair em LOA N+1 ou LOA N+2 a depender do mes/dia.
    Aqui marcamos o intervalo possivel.
    """
    return f"LOA {ano_autuacao+1} ou LOA {ano_autuacao+2}"

def cabe_em_2027(ano_autuacao: int) -> str:
    """LOA 2027 paga precatorios autuados entre 03/04/2025 e 02/04/2026.
    Ou seja, ano de autuacao 2025 (parte dela) ou 2026 (so ate 02/04)."""
    if ano_autuacao == 2025:
        return "POSSIVEL (autuacao 2025 — depende se foi apos 02/04/2025)"
    if ano_autuacao == 2026:
        return "POSSIVEL (autuacao 2026 — apenas se ate 02/04/2026)"
    if ano_autuacao == 2024:
        return "ANOMALIA — autuacao 2024 deveria pagar em LOA 2025 ou 2026, nao 2027"
    if ano_autuacao < 2024:
        return f"ANOMALIA — autuacao {ano_autuacao} muito antiga para LOA 2027"
    return f"ANOMALIA — autuacao {ano_autuacao} futura demais"

df["loa_prevista_pela_regra"] = df["ano_autuacao"].apply(loa_provavel)
df["status_lista_loa_2027"] = df["ano_autuacao"].apply(cabe_em_2027)

# 5) Tipo: precatorio vs RPV
# Esfera federal: limite RPV = 60 SM
# 2024: SM=R$1.412 -> RPV ate R$84.720
# 2025: SM=R$1.518 -> RPV ate R$91.080  (aprox; sera reajustado)
# 2026: SM=R$1.518 ainda valido para corte
def tipo_requisitorio(valor: float, ano: int) -> str:
    if ano == 2024:
        limite = 60 * 1412
    elif ano == 2025:
        limite = 60 * 1518
    else:
        limite = 60 * 1518
    if valor < limite:
        return f"RPV-PROVAVEL (valor < 60SM ano {ano} = R${limite:,.0f})"
    return "PRECATORIO (valor >= 60 SM)"

df["tipo_requisitorio"] = df.apply(lambda r: tipo_requisitorio(r["valor_brl"], r["ano_autuacao"]), axis=1)

# 6) Cruzar com LOA 2026 (precatorios_extraidos.csv) -- busca por substring de CNJ
# A LOA 2026 zera os 4 ultimos digitos (origem 0000). Comparar pelos primeiros 16.
def cnj_base_16(d19: str) -> str:
    return d19.zfill(20)[:16]

df["cnj_base_16"] = df["cnj_19digitos"].apply(cnj_base_16)

print("Carregando LOA 2026 (42k linhas)...")
loa = pd.read_csv(LOA2026, sep=";", dtype=str, encoding="utf-8")
def safe_base_16(x):
    s = str(x) if x is not None else ""
    s = s.replace(".0", "") if s.endswith(".0") else s
    if s.isdigit():
        return s.zfill(20)[:16]
    return ""

loa["cnj_base_16_loa"] = loa["Precatorio"].fillna("").astype(str).apply(safe_base_16)

merged = df.merge(
    loa.rename(columns={
        "Precatorio": "loa2026_cnj",
        "UO_Devedora_Nome": "loa2026_uo_devedora",
        "Tipo_Causa": "loa2026_tipo_causa",
        "Valor_RS": "loa2026_valor_centavos",
        "Ano": "loa2026_ano_orcamento",
    }),
    left_on="cnj_base_16",
    right_on="cnj_base_16_loa",
    how="left",
)

# Limpar colunas extras
merged = merged.drop(columns=["cnj_base_16_loa", "UO_Cadastradora_Codigo", "UO_Cadastradora_Nome",
                                "UO_Devedora_Codigo"], errors="ignore")

# Cruzar com SIOP conciliado v2 para informacoes adicionais
print("Carregando SIOP conciliado v2...")
siop = pd.read_csv(LOA2026_CONC, sep=";", dtype=str, encoding="latin-1")
siop["cnj_base_16_siop"] = siop["Precatorio"].fillna("").astype(str).apply(safe_base_16)

cols_siop = ["cnj_base_16_siop", "Valor Original do Precatório", "Valor Atualizado",
             "Data de Ajuizamento da Ação Originária", "Data da Autuação", "Fundef",
             "Class_Tempo", "FaixaValor", "status_conciliacao"]
siop_slim = siop[cols_siop].rename(columns={
    "Valor Original do Precatório": "siop_valor_original",
    "Valor Atualizado": "siop_valor_atualizado",
    "Data de Ajuizamento da Ação Originária": "siop_data_ajuizamento",
    "Data da Autuação": "siop_data_autuacao",
    "Fundef": "siop_fundef",
    "Class_Tempo": "siop_class_tempo",
    "FaixaValor": "siop_faixa_valor",
    "status_conciliacao": "siop_status_conciliacao",
})

merged = merged.merge(siop_slim, left_on="cnj_base_16", right_on="cnj_base_16_siop", how="left")
merged = merged.drop(columns=["cnj_base_16_siop", "cnj_base_16"], errors="ignore")

# Marcar se foi achado no LOA 2026
merged["match_loa2026"] = merged["loa2026_uo_devedora"].notna()
merged["match_siop"] = merged["siop_valor_original"].notna()

# Reordenar colunas
ordem_cols = [
    "ordem_oficio", "cnj_canonico", "cnj_19digitos",
    "ano_autuacao", "segmento", "tribunal", "origem",
    "valor_brl_str", "valor_brl",
    "tipo_requisitorio",
    "loa_prevista_pela_regra", "status_lista_loa_2027",
    "match_loa2026", "loa2026_ano_orcamento", "loa2026_uo_devedora", "loa2026_tipo_causa", "loa2026_valor_centavos",
    "match_siop", "siop_valor_original", "siop_valor_atualizado",
    "siop_data_ajuizamento", "siop_data_autuacao", "siop_fundef",
    "siop_class_tempo", "siop_faixa_valor", "siop_status_conciliacao",
]
merged = merged[[c for c in ordem_cols if c in merged.columns]]

# Salvar CSV
csv_out = f"{DEST}\\lista_loa2027_processada.csv"
merged.to_csv(csv_out, index=False, sep=";", encoding="utf-8-sig")
print(f"CSV salvo: {csv_out}")

# === XLSX ===
xlsx_out = f"{DEST}\\lista_loa2027_processada.xlsx"
with pd.ExcelWriter(xlsx_out, engine="openpyxl") as w:
    merged.to_excel(w, index=False, sheet_name="Lista LOA 2027")

    # Summary sheet
    summary = pd.DataFrame({
        "Indicador": [
            "Total de registros",
            "Total valor (R$)",
            "CNJs ano 2025 (encaixam em LOA 2027)",
            "CNJs ano 2024 (ANOMALIA na lista LOA 2027)",
            "Match com LOA 2026 (precatorios_extraidos)",
            "Match com SIOP conciliado v2",
            "Tipo PRECATORIO",
            "Tipo RPV-PROVAVEL",
            "TRF1 (origem 9198)",
        ],
        "Valor": [
            len(merged),
            f"R$ {merged['valor_brl'].sum():,.2f}",
            int((merged["ano_autuacao"] == 2025).sum()),
            int((merged["ano_autuacao"] == 2024).sum()),
            int(merged["match_loa2026"].sum()),
            int(merged["match_siop"].sum()),
            int((merged["tipo_requisitorio"].str.startswith("PRECATORIO")).sum()),
            int((merged["tipo_requisitorio"].str.startswith("RPV-PROVAVEL")).sum()),
            int((merged["tribunal"] == "TRF1").sum()),
        ]
    })
    summary.to_excel(w, index=False, sheet_name="Resumo")

# Formatacao
wb = load_workbook(xlsx_out)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1F4E78")
    align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", start_color="F2F2F2")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.font = Font(name="Arial", size=10)
            c.alignment = align
            c.border = border
            if row_idx % 2 == 0:
                c.fill = alt_fill
            v = ws.cell(row=row_idx, column=col_idx).value
            if isinstance(v, str) and v.startswith("ANOMALIA"):
                c.fill = PatternFill("solid", start_color="FFD9D9")
                c.font = Font(name="Arial", size=10, bold=True, color="C00000")

    # Larguras
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        header = ws.cell(row=1, column=col_idx).value or ""
        if "cnj" in header.lower():
            ws.column_dimensions[col_letter].width = 28
        elif "status" in header.lower() or "loa_prevista" in header.lower():
            ws.column_dimensions[col_letter].width = 50
        elif "uo_dev" in header.lower() or "tipo_causa" in header.lower():
            ws.column_dimensions[col_letter].width = 40
        elif "siop_data" in header.lower():
            ws.column_dimensions[col_letter].width = 22
        else:
            ws.column_dimensions[col_letter].width = 18

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"

    # Tabela
    last_col = get_column_letter(ws.max_column)
    tbl_name = f"Tab_{sheet_name.replace(' ', '_')}"
    tbl = Table(displayName=tbl_name, ref=f"A1:{last_col}{ws.max_row}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)

wb.save(xlsx_out)
print(f"XLSX salvo: {xlsx_out}")
print()
print("=" * 70)
print("RESUMO")
print("=" * 70)
print(f"Total registros: {len(merged)}")
print(f"Valor total: R$ {merged['valor_brl'].sum():,.2f}")
print(f"Ano 2025 (encaixa LOA 2027): {(merged['ano_autuacao']==2025).sum()}")
print(f"Ano 2024 (ANOMALIA): {(merged['ano_autuacao']==2024).sum()}")
print(f"Match LOA 2026: {merged['match_loa2026'].sum()}")
print(f"Match SIOP v2: {merged['match_siop'].sum()}")
print()
print(merged[["ordem_oficio", "cnj_canonico", "valor_brl", "status_lista_loa_2027",
              "match_loa2026", "loa2026_tipo_causa"]].to_string(index=False))
