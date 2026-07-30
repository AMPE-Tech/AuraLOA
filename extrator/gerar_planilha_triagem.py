"""
Gera a planilha de triagem de precatorios para selecao de pre-due-diligence.

Consome os JSON produzidos por extrator_precatorios.py e monta um workbook com:

  Triagem            uma linha por precatorio, com score e coluna de selecao
  Beneficiarios      uma linha por beneficiario (o dossie tipico tem varios)
  Valores            abertura de principal, juros, correcao, descontos
  Consistencia       conferencia aritmetica documento a documento
  Riscos_Pendencias  riscos, pendencias e conflitos de extracao
  Parametros         pesos e cortes — editaveis, alimentam as formulas
  Resumo             contagens e totais
  Dicionario         significado de cada coluna

As notas e o score sao FORMULAS, nao numeros gravados: mudar um peso na aba
Parametros reordena a triagem inteira sem reprocessar nenhum PDF.

Uso:
    python gerar_planilha_triagem.py saida/*.json -s triagem.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONTE = "Arial"

COR_CABECALHO = "1F3864"
COR_TEXTO_CABECALHO = "FFFFFF"
COR_ENTRADA = "FFF2CC"      # amarelo: preencher manualmente
COR_CALCULADO = "E8EFF7"    # azul claro: formula
COR_APROVAR = "C6EFCE"
COR_REVISAR = "FFEB9C"
COR_DESCARTAR = "FFC7CE"

FORMATO_REAL = '"R$" #,##0.00;("R$" #,##0.00);-'
FORMATO_PERCENTUAL = "0.0%"
FORMATO_NOTA = "0.0"

BORDA_FINA = Border(*(Side(style="thin", color="BFBFBF"),) * 4)


# --------------------------------------------------------------------------
# Leitura tolerante dos JSON
# --------------------------------------------------------------------------

def _valor(bruto: Any) -> float | None:
    """'R$ 3.523.707,62' -> 3523707.62 ; devolve None quando nao ha numero."""
    if bruto is None or bruto == "":
        return None
    if isinstance(bruto, (int, float)):
        return float(bruto)
    limpo = re.sub(r"[^\d,.-]", "", str(bruto))
    if not limpo:
        return None
    try:
        return float(limpo.replace(".", "").replace(",", "."))
    except ValueError:
        return None


def _percentual(bruto: Any) -> float | None:
    """Devolve fracao (0.40), que e como o Excel armazena porcentagem."""
    numero = _valor(str(bruto).replace("%", "")) if bruto not in (None, "") else None
    if numero is None:
        return None
    return numero / 100 if numero > 1 else numero


def _texto(bruto: Any, limite: int = 0) -> str:
    if bruto is None:
        return ""
    if isinstance(bruto, (list, tuple)):
        texto = " | ".join(str(item) for item in bruto if item)
    elif isinstance(bruto, dict):
        texto = " | ".join(f"{k}: {v}" for k, v in bruto.items() if v)
    else:
        texto = str(bruto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto[:limite] + "…" if limite and len(texto) > limite else texto


def _get(dados: dict, *caminho: str, padrao: Any = "") -> Any:
    atual: Any = dados
    for chave in caminho:
        if not isinstance(atual, dict):
            return padrao
        atual = atual.get(chave)
    return atual if atual not in (None, "") else padrao


# --------------------------------------------------------------------------
# Estilo
# --------------------------------------------------------------------------

def escrever_cabecalho(aba, colunas: list[tuple[str, int]], linha: int = 1) -> None:
    for indice, (titulo, largura) in enumerate(colunas, 1):
        celula = aba.cell(row=linha, column=indice, value=titulo)
        celula.font = Font(name=FONTE, bold=True, size=10, color=COR_TEXTO_CABECALHO)
        celula.fill = PatternFill("solid", fgColor=COR_CABECALHO)
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celula.border = BORDA_FINA
        aba.column_dimensions[get_column_letter(indice)].width = largura
    aba.row_dimensions[linha].height = 34
    aba.freeze_panes = aba.cell(row=linha + 1, column=1)


def estilizar_corpo(aba, primeira_linha: int, ultima_linha: int, ultima_coluna: int) -> None:
    for linha in range(primeira_linha, ultima_linha + 1):
        for coluna in range(1, ultima_coluna + 1):
            celula = aba.cell(row=linha, column=coluna)
            celula.font = Font(name=FONTE, size=10)
            celula.border = BORDA_FINA
            celula.alignment = Alignment(vertical="top", wrap_text=True)


# --------------------------------------------------------------------------
# Aba Parametros
# --------------------------------------------------------------------------

PARAMETROS = [
    ("Peso — Valor do credito", 30, "Quanto o tamanho do credito pesa no score (0 a 100)."),
    ("Peso — Maturidade documental", 25, "Transito em julgado, precatorio autuado, ausencia de pendencias."),
    ("Peso — Preferencia", 15, "Natureza alimenticia e superpreferencia (idoso, doenca grave, PcD)."),
    ("Peso — Liquidez", 15, "Existencia de janela de acordo direto e tamanho do desagio."),
    ("Peso — Completude da extracao", 15, "Cobertura dos campos essenciais no documento lido."),
    ("Valor de referencia p/ nota 10 (R$)", 5000000, "Credito neste valor recebe nota 10 na dimensao Valor."),
    ("Corte — APROVAR (score >=)", 70, "Score igual ou acima classifica como APROVAR."),
    ("Corte — REVISAR (score >=)", 45, "Entre este corte e o de APROVAR classifica como REVISAR."),
    ("Penalidade por pendencia documental", 0.5, "Descontado da nota de maturidade, ate o teto de 3 pontos."),
    ("Penalidade — sem transito em julgado", 4, "Descontado da nota de maturidade."),
    ("Penalidade — sem numero de precatorio", 3, "Descontado da nota de maturidade."),
]


def montar_parametros(planilha: Workbook):
    aba = planilha.create_sheet("Parametros")
    aba["A1"] = "PARAMETROS DE TRIAGEM"
    aba["A1"].font = Font(name=FONTE, bold=True, size=13, color=COR_CABECALHO)
    aba["A2"] = (
        "Edite apenas as celulas amarelas da coluna B. Todas as notas e o score da aba "
        "Triagem sao formulas que apontam para ca — alterar um peso reordena a triagem "
        "inteira sem reprocessar nenhum PDF. A soma dos cinco pesos deve dar 100."
    )
    aba["A2"].font = Font(name=FONTE, size=9, italic=True)
    aba["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    aba.merge_cells("A2:C2")
    aba.row_dimensions[2].height = 42

    escrever_cabecalho(aba, [("Parametro", 40), ("Valor", 16), ("O que faz", 68)], linha=4)

    for deslocamento, (nome, valor, descricao) in enumerate(PARAMETROS):
        linha = 5 + deslocamento
        aba.cell(row=linha, column=1, value=nome)
        celula = aba.cell(row=linha, column=2, value=valor)
        celula.fill = PatternFill("solid", fgColor=COR_ENTRADA)
        celula.font = Font(name=FONTE, size=10, bold=True, color="0000FF")
        if "R$" in nome:
            celula.number_format = FORMATO_REAL
        aba.cell(row=linha, column=3, value=descricao)

    linha_soma = 5 + len(PARAMETROS) + 1
    aba.cell(row=linha_soma, column=1, value="Soma dos pesos (deve ser 100)").font = Font(
        name=FONTE, size=10, bold=True
    )
    aba.cell(row=linha_soma, column=2, value="=SUM(B5:B9)").font = Font(name=FONTE, size=10, bold=True)
    aba.cell(row=linha_soma, column=3, value="Se nao fechar 100, o score deixa de ser comparavel entre linhas.")

    estilizar_corpo(aba, 5, linha_soma, 3)
    aba.freeze_panes = "A5"
    return aba


# Enderecos fixos usados pelas formulas da aba Triagem.
P_PESO_VALOR = "Parametros!$B$5"
P_PESO_MATURIDADE = "Parametros!$B$6"
P_PESO_PREFERENCIA = "Parametros!$B$7"
P_PESO_LIQUIDEZ = "Parametros!$B$8"
P_PESO_COMPLETUDE = "Parametros!$B$9"
P_VALOR_REFERENCIA = "Parametros!$B$10"
P_CORTE_APROVAR = "Parametros!$B$11"
P_CORTE_REVISAR = "Parametros!$B$12"
P_PEN_PENDENCIA = "Parametros!$B$13"
P_PEN_TRANSITO = "Parametros!$B$14"
P_PEN_PRECATORIO = "Parametros!$B$15"


# --------------------------------------------------------------------------
# Aba Triagem
# --------------------------------------------------------------------------

COLUNAS_TRIAGEM = [
    ("ID", 6), ("Nº Precatório", 16), ("Processo CNJ", 24), ("Tribunal", 14),
    ("Ente Devedor", 26), ("Natureza", 14), ("Beneficiário Principal", 34),
    ("Qtd. Benef.", 9), ("Data-Base", 12), ("Trânsito em Julgado", 13),
    ("Valor Bruto Requisitado", 18), ("Valor Atualizado", 18), ("Deságio", 10),
    ("Valor Líquido p/ Acordo", 18), ("Prazo de Adesão", 13),
    ("Super-preferência", 11), ("Pendências", 10), ("Riscos", 8),
    ("Cobertura Extração", 11),
    ("Nota Valor", 9), ("Nota Maturidade", 10), ("Nota Preferência", 10),
    ("Nota Liquidez", 9), ("Nota Completude", 10),
    ("SCORE", 10), ("Classificação Sugerida", 15),
    ("SELECIONAR p/ Pré-DD", 15), ("Responsável", 16), ("Observações", 42),
]

# indices de coluna (1-based) reutilizados nas formulas
C_PRECATORIO, C_NATUREZA = 2, 6
C_TRANSITO = 10
C_BRUTO, C_ATUALIZADO, C_DESAGIO, C_LIQUIDO = 11, 12, 13, 14
C_PRAZO, C_SUPERPREF, C_PENDENCIAS, C_RISCOS, C_COBERTURA = 15, 16, 17, 18, 19
C_NOTA_VALOR, C_NOTA_MAT, C_NOTA_PREF, C_NOTA_LIQ, C_NOTA_COMP = 20, 21, 22, 23, 24
C_SCORE, C_CLASSIFICACAO, C_SELECAO, C_RESPONSAVEL, C_OBSERVACOES = 25, 26, 27, 28, 29


def montar_triagem(planilha: Workbook, documentos: list[dict]) -> None:
    aba = planilha.create_sheet("Triagem", 0)
    aba["A1"] = "TRIAGEM DE PRECATÓRIOS — SELEÇÃO PARA PRÉ-DUE-DILIGENCE"
    aba["A1"].font = Font(name=FONTE, bold=True, size=14, color=COR_CABECALHO)
    aba["A2"] = (
        "Colunas azuis são calculadas por fórmula — não digite nelas. Colunas amarelas "
        "(SELECIONAR, Responsável, Observações) são suas. A classificação é uma sugestão "
        "do score, não uma decisão: a coluna SELECIONAR é que vale."
    )
    aba["A2"].font = Font(name=FONTE, size=9, italic=True)
    aba["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    aba.merge_cells("A2:H2")
    aba.row_dimensions[2].height = 30

    escrever_cabecalho(aba, COLUNAS_TRIAGEM, linha=4)
    primeira = 5

    for deslocamento, documento in enumerate(documentos):
        linha = primeira + deslocamento
        processuais = documento.get("dados_processuais", {})
        beneficiarios = documento.get("beneficiarios", []) or []
        acordo = documento.get("acordo_direto", {})
        controle = documento.get("controle_extracao", {})
        superpreferencia = documento.get("superpreferencia", {}) or {}

        atualizado = sum(
            _valor(b.get("valor_atualizado")) or 0.0 for b in beneficiarios
        ) or _valor(_get(documento, "valores", "valor_bruto_requisicao"))

        principal = next(
            (b for b in beneficiarios if str(b.get("papel", "")).lower().startswith("titular")),
            beneficiarios[0] if beneficiarios else {},
        )

        aba.cell(row=linha, column=1, value=deslocamento + 1)
        aba.cell(row=linha, column=C_PRECATORIO, value=_texto(processuais.get("numero_precatorio")))
        aba.cell(row=linha, column=3, value=_texto(processuais.get("numero_processo_cnj")))
        aba.cell(row=linha, column=4, value=_texto(processuais.get("tribunal"), 40))
        aba.cell(row=linha, column=5, value=_texto(_get(documento, "partes", "devedor_executado"), 60))
        aba.cell(row=linha, column=C_NATUREZA, value=_texto(processuais.get("natureza_credito")))
        aba.cell(row=linha, column=7, value=_texto(principal.get("nome") or
                                                   _get(documento, "partes", "credor_beneficiario"), 60))
        aba.cell(row=linha, column=8, value=len(beneficiarios))
        aba.cell(row=linha, column=9, value=_texto(processuais.get("data_base_calculo")))
        aba.cell(row=linha, column=C_TRANSITO, value=_texto(processuais.get("data_transito_julgado")))

        celula = aba.cell(row=linha, column=C_BRUTO,
                          value=_valor(_get(documento, "valores", "valor_bruto_requisicao")))
        celula.number_format = FORMATO_REAL
        celula = aba.cell(row=linha, column=C_ATUALIZADO, value=atualizado)
        celula.number_format = FORMATO_REAL
        celula = aba.cell(row=linha, column=C_DESAGIO, value=_percentual(acordo.get("desagio_percentual")))
        celula.number_format = FORMATO_PERCENTUAL

        celula = aba.cell(
            row=linha, column=C_LIQUIDO,
            value=f"=IF(OR({_ref(C_ATUALIZADO, linha)}=\"\",{_ref(C_DESAGIO, linha)}=\"\"),\"\","
                  f"{_ref(C_ATUALIZADO, linha)}*(1-{_ref(C_DESAGIO, linha)}))",
        )
        celula.number_format = FORMATO_REAL
        celula.fill = PatternFill("solid", fgColor=COR_CALCULADO)

        aba.cell(row=linha, column=C_PRAZO, value=_texto(acordo.get("prazo_adesao")))
        aba.cell(row=linha, column=C_SUPERPREF,
                 value="SIM" if any(superpreferencia.values()) else "NÃO")
        aba.cell(row=linha, column=C_PENDENCIAS,
                 value=len(documento.get("dados_ausentes_ou_duvidosos", []) or []))
        aba.cell(row=linha, column=C_RISCOS, value=len(documento.get("riscos", []) or []))
        celula = aba.cell(row=linha, column=C_COBERTURA,
                          value=controle.get("cobertura_campos_essenciais_num", 0))
        celula.number_format = FORMATO_PERCENTUAL

        formulas = {
            C_NOTA_VALOR:
                f"=IF({_ref(C_ATUALIZADO, linha)}=\"\",0,"
                f"MIN(10,{_ref(C_ATUALIZADO, linha)}/{P_VALOR_REFERENCIA}*10))",
            C_NOTA_MAT:
                f"=MAX(0,10-IF({_ref(C_TRANSITO, linha)}=\"\",{P_PEN_TRANSITO},0)"
                f"-IF({_ref(C_PRECATORIO, linha)}=\"\",{P_PEN_PRECATORIO},0)"
                f"-MIN(3,{_ref(C_PENDENCIAS, linha)}*{P_PEN_PENDENCIA}))",
            C_NOTA_PREF:
                f"=MIN(10,IF(ISNUMBER(SEARCH(\"ALIMENT\",{_ref(C_NATUREZA, linha)})),7,3)"
                f"+IF({_ref(C_SUPERPREF, linha)}=\"SIM\",3,0))",
            C_NOTA_LIQ:
                f"=IF(AND({_ref(C_PRAZO, linha)}<>\"\",{_ref(C_DESAGIO, linha)}<>\"\"),"
                f"MAX(0,10-{_ref(C_DESAGIO, linha)}*10),"
                f"IF({_ref(C_PRAZO, linha)}<>\"\",6,4))",
            C_NOTA_COMP: f"={_ref(C_COBERTURA, linha)}*10",
            C_SCORE:
                f"=({_ref(C_NOTA_VALOR, linha)}*{P_PESO_VALOR}"
                f"+{_ref(C_NOTA_MAT, linha)}*{P_PESO_MATURIDADE}"
                f"+{_ref(C_NOTA_PREF, linha)}*{P_PESO_PREFERENCIA}"
                f"+{_ref(C_NOTA_LIQ, linha)}*{P_PESO_LIQUIDEZ}"
                f"+{_ref(C_NOTA_COMP, linha)}*{P_PESO_COMPLETUDE})/10",
            C_CLASSIFICACAO:
                f"=IF({_ref(C_SCORE, linha)}>={P_CORTE_APROVAR},\"APROVAR\","
                f"IF({_ref(C_SCORE, linha)}>={P_CORTE_REVISAR},\"REVISAR\",\"DESCARTAR\"))",
        }
        for coluna, formula in formulas.items():
            celula = aba.cell(row=linha, column=coluna, value=formula)
            celula.fill = PatternFill("solid", fgColor=COR_CALCULADO)
            celula.number_format = FORMATO_NOTA if coluna != C_CLASSIFICACAO else "General"
            celula.alignment = Alignment(horizontal="center", vertical="center")
            if coluna == C_SCORE:
                celula.font = Font(name=FONTE, size=10, bold=True)

        for coluna in (C_SELECAO, C_RESPONSAVEL, C_OBSERVACOES):
            celula = aba.cell(row=linha, column=coluna)
            celula.fill = PatternFill("solid", fgColor=COR_ENTRADA)
        aba.cell(row=linha, column=C_SELECAO, value="PENDENTE").alignment = Alignment(
            horizontal="center"
        )

    ultima = primeira + max(len(documentos), 1) - 1
    estilizar_corpo(aba, primeira, ultima, len(COLUNAS_TRIAGEM))
    aba.freeze_panes = "C5"

    # Reaplica formatos que estilizar_corpo sobrescreve.
    for linha in range(primeira, ultima + 1):
        aba.cell(row=linha, column=C_SCORE).font = Font(name=FONTE, size=10, bold=True)
        for coluna in (C_NOTA_VALOR, C_NOTA_MAT, C_NOTA_PREF, C_NOTA_LIQ, C_NOTA_COMP,
                       C_SCORE, C_CLASSIFICACAO, C_SELECAO):
            aba.cell(row=linha, column=coluna).alignment = Alignment(
                horizontal="center", vertical="center"
            )

    faixa_selecao = f"{_letra(C_SELECAO)}{primeira}:{_letra(C_SELECAO)}{ultima + 200}"
    validacao = DataValidation(
        type="list", formula1='"SIM,NÃO,PENDENTE"', allow_blank=True, showDropDown=False
    )
    validacao.error = "Escolha SIM, NÃO ou PENDENTE."
    validacao.prompt = "Este precatório segue para pré-due-diligence?"
    aba.add_data_validation(validacao)
    validacao.add(faixa_selecao)

    faixa_classificacao = f"{_letra(C_CLASSIFICACAO)}{primeira}:{_letra(C_CLASSIFICACAO)}{ultima}"
    for texto, cor in (("APROVAR", COR_APROVAR), ("REVISAR", COR_REVISAR), ("DESCARTAR", COR_DESCARTAR)):
        aba.conditional_formatting.add(
            faixa_classificacao,
            CellIsRule(operator="equal", formula=[f'"{texto}"'],
                       fill=PatternFill("solid", bgColor=cor)),
        )
    aba.conditional_formatting.add(
        faixa_selecao,
        CellIsRule(operator="equal", formula=['"SIM"'],
                   fill=PatternFill("solid", bgColor=COR_APROVAR)),
    )
    aba.auto_filter.ref = f"A4:{_letra(len(COLUNAS_TRIAGEM))}{ultima}"


def _letra(coluna: int) -> str:
    return get_column_letter(coluna)


def _ref(coluna: int, linha: int) -> str:
    return f"${_letra(coluna)}{linha}"


# --------------------------------------------------------------------------
# Abas de detalhe
# --------------------------------------------------------------------------

def montar_beneficiarios(planilha: Workbook, documentos: list[dict]) -> None:
    aba = planilha.create_sheet("Beneficiarios")
    colunas = [
        ("Nº Precatório", 16), ("Beneficiário", 38), ("Papel", 18), ("CPF/CNPJ", 18),
        ("OAB", 12), ("%", 8), ("Principal (data-base)", 17), ("Juros (data-base)", 17),
        ("Total (data-base)", 17), ("Valor Atualizado", 17), ("Deságio", 10),
        ("Líquido p/ Acordo", 17), ("Banco", 24), ("Agência", 11), ("Conta", 13),
    ]
    escrever_cabecalho(aba, colunas)
    linha = 2
    for documento in documentos:
        precatorio = _texto(_get(documento, "dados_processuais", "numero_precatorio"))
        for beneficiario in documento.get("beneficiarios", []) or []:
            aba.cell(row=linha, column=1, value=precatorio)
            aba.cell(row=linha, column=2, value=_texto(beneficiario.get("nome")))
            aba.cell(row=linha, column=3, value=_texto(beneficiario.get("papel")))
            aba.cell(row=linha, column=4, value=_texto(beneficiario.get("documento")))
            aba.cell(row=linha, column=5, value=_texto(beneficiario.get("oab")))
            celula = aba.cell(row=linha, column=6, value=_percentual(beneficiario.get("percentual")))
            celula.number_format = FORMATO_PERCENTUAL
            for deslocamento, chave in enumerate(
                ["valor_principal", "valor_juros", "valor_total", "valor_atualizado"]
            ):
                celula = aba.cell(row=linha, column=7 + deslocamento, value=_valor(beneficiario.get(chave)))
                celula.number_format = FORMATO_REAL
            celula = aba.cell(row=linha, column=11,
                              value=_percentual(beneficiario.get("desagio_percentual")))
            celula.number_format = FORMATO_PERCENTUAL
            celula = aba.cell(
                row=linha, column=12,
                value=f"=IF(OR(J{linha}=\"\",K{linha}=\"\"),\"\",J{linha}*(1-K{linha}))",
            )
            celula.number_format = FORMATO_REAL
            celula.fill = PatternFill("solid", fgColor=COR_CALCULADO)
            aba.cell(row=linha, column=13, value=_texto(beneficiario.get("banco")))
            aba.cell(row=linha, column=14, value=_texto(beneficiario.get("agencia")))
            aba.cell(row=linha, column=15, value=_texto(beneficiario.get("conta")))
            linha += 1

    if linha > 2:
        aba.cell(row=linha, column=2, value="TOTAL").font = Font(name=FONTE, bold=True, size=10)
        for coluna in (7, 8, 9, 10, 12):
            celula = aba.cell(
                row=linha, column=coluna,
                value=f"=SUM({_letra(coluna)}2:{_letra(coluna)}{linha - 1})",
            )
            celula.number_format = FORMATO_REAL
            celula.font = Font(name=FONTE, bold=True, size=10)
    estilizar_corpo(aba, 2, max(linha, 2), len(colunas))


def montar_valores(planilha: Workbook, documentos: list[dict]) -> None:
    aba = planilha.create_sheet("Valores")
    colunas = [
        ("Nº Precatório", 16), ("Valor Bruto Requisição", 18), ("Principal Corrigido", 18),
        ("Juros", 18), ("Honorários", 18), ("Desconto Previdenciário", 18),
        ("Descontos/Retenções", 30), ("Correção Monetária", 46), ("Observações", 60),
    ]
    escrever_cabecalho(aba, colunas)
    linha = 2
    for documento in documentos:
        valores = documento.get("valores", {})
        aba.cell(row=linha, column=1,
                 value=_texto(_get(documento, "dados_processuais", "numero_precatorio")))
        for deslocamento, chave in enumerate(
            ["valor_bruto_requisicao", "valor_principal", "juros", "honorarios",
             "desconto_previdenciario"]
        ):
            celula = aba.cell(row=linha, column=2 + deslocamento, value=_valor(valores.get(chave)))
            celula.number_format = FORMATO_REAL
        aba.cell(row=linha, column=7, value=_texto(valores.get("descontos_retencoes"), 300))
        aba.cell(row=linha, column=8, value=_texto(valores.get("correcao_monetaria"), 400))
        aba.cell(row=linha, column=9, value=_texto(valores.get("observacoes_valores"), 900))
        linha += 1
    estilizar_corpo(aba, 2, max(linha - 1, 2), len(colunas))


def montar_consistencia(planilha: Workbook, documentos: list[dict]) -> None:
    aba = planilha.create_sheet("Consistencia")
    colunas = [
        ("Nº Precatório", 16), ("Checagem", 40), ("Fórmula conferida", 46),
        ("Situação", 14), ("Diferença (R$)", 15), ("Observação", 34),
    ]
    escrever_cabecalho(aba, colunas)
    linha = 2
    for documento in documentos:
        precatorio = _texto(_get(documento, "dados_processuais", "numero_precatorio"))
        for checagem in documento.get("consistencia", []) or []:
            aba.cell(row=linha, column=1, value=precatorio)
            aba.cell(row=linha, column=2, value=_texto(checagem.get("checagem")))
            aba.cell(row=linha, column=3, value=_texto(checagem.get("formula")))
            aba.cell(row=linha, column=4, value=_texto(checagem.get("situacao")))
            celula = aba.cell(row=linha, column=5, value=_valor(checagem.get("diferenca")))
            celula.number_format = FORMATO_REAL
            aba.cell(row=linha, column=6, value=_texto(checagem.get("observacao")))
            linha += 1
    estilizar_corpo(aba, 2, max(linha - 1, 2), len(colunas))
    if linha > 2:
        for texto, cor in (("OK", COR_APROVAR), ("DIVERGENTE", COR_DESCARTAR),
                           ("NAO_APLICAVEL", COR_REVISAR)):
            aba.conditional_formatting.add(
                f"D2:D{linha - 1}",
                CellIsRule(operator="equal", formula=[f'"{texto}"'],
                           fill=PatternFill("solid", bgColor=cor)),
            )


def montar_riscos(planilha: Workbook, documentos: list[dict]) -> None:
    aba = planilha.create_sheet("Riscos_Pendencias")
    colunas = [("Nº Precatório", 16), ("Categoria", 20), ("Item", 110), ("Tratado?", 12)]
    escrever_cabecalho(aba, colunas)
    linha = 2
    categorias = [
        ("Risco", "riscos"),
        ("Pendência / dúvida", "dados_ausentes_ou_duvidosos"),
        ("Prazo", "prazos"),
        ("Obrigação", "obrigacoes"),
        ("Oportunidade", "oportunidades"),
        ("Decisão necessária", "decisoes_necessarias"),
        ("Ação sugerida", "acoes_sugeridas"),
    ]
    for documento in documentos:
        precatorio = _texto(_get(documento, "dados_processuais", "numero_precatorio"))
        for rotulo, chave in categorias:
            for item in documento.get(chave, []) or []:
                aba.cell(row=linha, column=1, value=precatorio)
                aba.cell(row=linha, column=2, value=rotulo)
                aba.cell(row=linha, column=3, value=_texto(item, 600))
                celula = aba.cell(row=linha, column=4, value="NÃO")
                celula.fill = PatternFill("solid", fgColor=COR_ENTRADA)
                celula.alignment = Alignment(horizontal="center")
                linha += 1
        for conflito in documento.get("conflitos_detectados", []) or []:
            aba.cell(row=linha, column=1, value=precatorio)
            aba.cell(row=linha, column=2, value="Conflito de extração")
            aba.cell(
                row=linha, column=3,
                value=f"{conflito.get('campo')} — escolhido: {_texto(conflito.get('escolhido'), 120)}"
                      f" | alternativas: {_texto(conflito.get('alternativas'), 300)}",
            )
            celula = aba.cell(row=linha, column=4, value="NÃO")
            celula.fill = PatternFill("solid", fgColor=COR_ENTRADA)
            linha += 1

    estilizar_corpo(aba, 2, max(linha - 1, 2), len(colunas))
    if linha > 2:
        validacao = DataValidation(type="list", formula1='"SIM,NÃO,N/A"', allow_blank=True)
        aba.add_data_validation(validacao)
        validacao.add(f"D2:D{linha + 200}")


def montar_resumo(planilha: Workbook, total_linhas: int) -> None:
    aba = planilha.create_sheet("Resumo")
    aba["A1"] = "RESUMO DA TRIAGEM"
    aba["A1"].font = Font(name=FONTE, bold=True, size=14, color=COR_CABECALHO)
    aba.column_dimensions["A"].width = 44
    aba.column_dimensions["B"].width = 22

    primeira, ultima = 5, 4 + max(total_linhas, 1)
    faixa_classificacao = f"Triagem!{_letra(C_CLASSIFICACAO)}{primeira}:{_letra(C_CLASSIFICACAO)}{ultima}"
    faixa_selecao = f"Triagem!{_letra(C_SELECAO)}{primeira}:{_letra(C_SELECAO)}{ultima}"
    faixa_atualizado = f"Triagem!{_letra(C_ATUALIZADO)}{primeira}:{_letra(C_ATUALIZADO)}{ultima}"
    faixa_liquido = f"Triagem!{_letra(C_LIQUIDO)}{primeira}:{_letra(C_LIQUIDO)}{ultima}"

    itens = [
        ("Precatórios na base", f"=COUNTA(Triagem!$C${primeira}:$C${ultima})", "0"),
        ("Classificados APROVAR", f'=COUNTIF({faixa_classificacao},"APROVAR")', "0"),
        ("Classificados REVISAR", f'=COUNTIF({faixa_classificacao},"REVISAR")', "0"),
        ("Classificados DESCARTAR", f'=COUNTIF({faixa_classificacao},"DESCARTAR")', "0"),
        ("", "", ""),
        ("Selecionados p/ pré-DD (SIM)", f'=COUNTIF({faixa_selecao},"SIM")', "0"),
        ("Recusados (NÃO)", f'=COUNTIF({faixa_selecao},"NÃO")', "0"),
        ("Ainda pendentes", f'=COUNTIF({faixa_selecao},"PENDENTE")', "0"),
        ("", "", ""),
        ("Valor atualizado — carteira total", f"=SUM({faixa_atualizado})", FORMATO_REAL),
        ("Valor líquido — carteira total", f"=SUM({faixa_liquido})", FORMATO_REAL),
        ("Valor atualizado — selecionados", f'=SUMIF({faixa_selecao},"SIM",{faixa_atualizado})', FORMATO_REAL),
        ("Valor líquido — selecionados", f'=SUMIF({faixa_selecao},"SIM",{faixa_liquido})', FORMATO_REAL),
        ("", "", ""),
        ("Score médio da base", f"=IFERROR(AVERAGE(Triagem!${_letra(C_SCORE)}${primeira}:"
                               f"${_letra(C_SCORE)}${ultima}),0)", "0.0"),
        ("Maior score", f"=IFERROR(MAX(Triagem!${_letra(C_SCORE)}${primeira}:"
                        f"${_letra(C_SCORE)}${ultima}),0)", "0.0"),
    ]

    linha = 3
    for rotulo, formula, formato in itens:
        if not rotulo:
            linha += 1
            continue
        celula_rotulo = aba.cell(row=linha, column=1, value=rotulo)
        celula_rotulo.font = Font(name=FONTE, size=10, bold=rotulo.startswith("Valor"))
        celula = aba.cell(row=linha, column=2, value=formula)
        celula.number_format = formato
        celula.font = Font(name=FONTE, size=10, bold=True)
        celula.fill = PatternFill("solid", fgColor=COR_CALCULADO)
        celula.border = BORDA_FINA
        linha += 1


DICIONARIO = [
    ("Valor Bruto Requisitado", "Triagem", "Valor do ofício requisitório na data-base, antes de atualização."),
    ("Valor Atualizado", "Triagem", "Soma dos valores atualizados de todos os beneficiários na última data de cálculo do documento."),
    ("Deságio", "Triagem", "Percentual de desconto do edital de acordo direto, quando houver."),
    ("Valor Líquido p/ Acordo", "Triagem", "Fórmula: Valor Atualizado × (1 − Deságio)."),
    ("Cobertura Extração", "Triagem", "Percentual dos 16 campos essenciais localizados no documento. Abaixo de 70% o registro merece releitura."),
    ("Nota Valor", "Triagem", "MIN(10; Valor Atualizado ÷ valor de referência × 10)."),
    ("Nota Maturidade", "Triagem", "Parte de 10 e desconta ausência de trânsito em julgado, ausência de número de precatório e pendências."),
    ("Nota Preferência", "Triagem", "7 se alimentícia, 3 se comum; +3 com superpreferência."),
    ("Nota Liquidez", "Triagem", "10 − (deságio × 10) quando há janela de acordo; 6 com edital sem prazo; 4 sem janela."),
    ("Nota Completude", "Triagem", "Cobertura da extração × 10."),
    ("SCORE", "Triagem", "Média das cinco notas ponderada pelos pesos da aba Parâmetros, em escala de 0 a 100."),
    ("Classificação Sugerida", "Triagem", "APROVAR, REVISAR ou DESCARTAR conforme os cortes da aba Parâmetros. É sugestão, não decisão."),
    ("SELECIONAR p/ Pré-DD", "Triagem", "SUA decisão: SIM, NÃO ou PENDENTE. Alimenta os totais da aba Resumo."),
    ("Situação", "Consistencia", "OK = a conta do documento fecha. DIVERGENTE = confira antes de avançar. NAO_APLICAVEL = faltou dado."),
    ("Conflito de extração", "Riscos_Pendencias", "Campo com mais de um valor candidato no documento. Mostra o escolhido e as alternativas."),
]


def montar_dicionario(planilha: Workbook) -> None:
    aba = planilha.create_sheet("Dicionario")
    escrever_cabecalho(aba, [("Campo", 28), ("Aba", 20), ("Definição", 96)])
    for deslocamento, (campo, origem, definicao) in enumerate(DICIONARIO):
        linha = 2 + deslocamento
        aba.cell(row=linha, column=1, value=campo).font = Font(name=FONTE, size=10, bold=True)
        aba.cell(row=linha, column=2, value=origem)
        aba.cell(row=linha, column=3, value=definicao)
    estilizar_corpo(aba, 2, 1 + len(DICIONARIO), 3)


# --------------------------------------------------------------------------
# Montagem
# --------------------------------------------------------------------------

def gerar_planilha(documentos: list[dict], destino: Path) -> Path:
    planilha = Workbook()
    planilha.remove(planilha.active)

    montar_triagem(planilha, documentos)
    montar_parametros(planilha)
    montar_beneficiarios(planilha, documentos)
    montar_valores(planilha, documentos)
    montar_consistencia(planilha, documentos)
    montar_riscos(planilha, documentos)
    montar_resumo(planilha, len(documentos))
    montar_dicionario(planilha)

    # A aba Triagem ja nasce no indice 0 via create_sheet(..., 0); nao ha
    # reordenacao a fazer aqui.
    destino.parent.mkdir(parents=True, exist_ok=True)
    planilha.save(destino)
    return destino


def main() -> None:
    analisador = argparse.ArgumentParser(description="Monta a planilha de triagem.")
    analisador.add_argument("jsons", nargs="+", type=Path)
    analisador.add_argument("-s", "--saida", type=Path, default=Path("triagem_precatorios.xlsx"))
    argumentos = analisador.parse_args()

    documentos = []
    for caminho in argumentos.jsons:
        try:
            documentos.append(json.loads(caminho.read_text(encoding="utf-8")))
        except Exception as erro:  # noqa: BLE001
            print(f"[{caminho.name}] ignorado: {erro}", file=sys.stderr)

    if not documentos:
        print("Nenhum JSON valido.", file=sys.stderr)
        raise SystemExit(1)

    destino = gerar_planilha(documentos, argumentos.saida)
    print(f"Planilha gerada: {destino} ({len(documentos)} precatorios)")


if __name__ == "__main__":
    main()
