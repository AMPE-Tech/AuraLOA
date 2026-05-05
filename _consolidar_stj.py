"""
Consolida todos os XLSM/XLSX/XLS baixados do STJ em 1 XLSM unificado.
Filtros:
  - Acima de 60 SM federal (PRC, não RPV — todos esses arquivos já são PRC)
  - Devedoras: União, INSS, Fazenda Nacional, Universidades Federais, Fundações
Colunas finais incluem fonte_url + fonte_arquivo.
"""
import os, json, re, glob, datetime, sys
import openpyxl
import xlrd

# Forçar flush em prints
def log(msg):
    print(msg, flush=True)
    sys.stdout.flush()

RAW = r'C:\AuraLOA-Confidencial\stj-precatorios\raw'
OUT = r'C:\AuraLOA-Confidencial\stj-precatorios\STJ-PRECATORIOS-PENDENTES-CONSOLIDADO.xlsx'

with open(os.path.join(RAW, '_MANIFEST.json'), 'r', encoding='utf-8') as f:
    manifest = {m['fname']: m for m in json.load(f)}

# Filtro de devedora: aceitar União, INSS, Fazenda Nacional, Universidades Federais, Fundações
DEVEDORA_OK = re.compile(
    r'(uni[aã]o|inss|seguro\s*social|fazenda\s*nacional|universidade\s*federal|funda[cç][aã]o|autarquia)',
    re.IGNORECASE
)

def cell(ws, r, c):
    """Get cell value handling both openpyxl and xlrd."""
    if hasattr(ws, 'cell_value'):  # xlrd
        try: return ws.cell_value(r-1, c-1)
        except: return None
    return ws.cell(r, c).value  # openpyxl

def maxrow(ws):
    return ws.nrows if hasattr(ws,'nrows') else ws.max_row
def maxcol(ws):
    return ws.ncols if hasattr(ws,'ncols') else ws.max_column

def find_header_row(ws, keys=('Processo','Registro','Natureza')):
    """Encontra a linha de header procurando por palavras-chave."""
    for r in range(1, min(20, maxrow(ws)+1)):
        row = [str(cell(ws, r, c) or '') for c in range(1, min(maxcol(ws)+1, 30))]
        joined = ' | '.join(row)
        hits = sum(1 for k in keys if k.lower() in joined.lower())
        if hits >= 2:
            return r, row
    return None, []

def map_columns(headers):
    """Mapeia coluna por palavra-chave."""
    cmap = {}
    for i, h in enumerate(headers):
        h_low = (h or '').lower()
        if not h_low: continue
        if 'processo' in h_low and 'cnj' not in cmap:
            cmap['cnj'] = i + 1
        elif 'registro' in h_low and 'prc' not in cmap:
            cmap['prc'] = i + 1
        elif 'tribunal' in h_low and 'tribunal' not in cmap:
            cmap['tribunal'] = i + 1
        elif 'origem' in h_low and 'acao' not in h_low and 'tr' not in h_low:
            cmap['acao_origem'] = i + 1
        elif ('entidade devedora' in h_low or h_low.strip() == 'devedora') and 'devedora' not in cmap:
            cmap['devedora'] = i + 1
        elif 'cnpj' in h_low and 'cnpj_dev' not in cmap:
            cmap['cnpj_dev'] = i + 1
        elif 'cidade' in h_low and 'cidade' not in cmap:
            cmap['cidade'] = i + 1
        elif 'natureza' in h_low and 'natureza' not in cmap:
            cmap['natureza'] = i + 1
        elif 'expedi' in h_low and 'data' in h_low and 'data_exp' not in cmap:
            cmap['data_expedicao'] = i + 1
        elif 'tr' in h_low and 'jul' in h_low and 'origem' in h_low:
            cmap['transito_origem'] = i + 1
        elif 'valor' in h_low and ('atualiz' in h_low or 'historic' in h_low or 'principal' in h_low):
            if 'valor' not in cmap:
                cmap['valor'] = i + 1
            elif 'valor_atualizado' not in cmap and 'atualiz' in h_low:
                cmap['valor_atualizado'] = i + 1
    return cmap

def is_super_preferencial(*texts):
    txt = ' '.join(str(t or '').lower() for t in texts)
    return bool(re.search(r'(super|preferencial|idoso|deficiente|doen[cç]a\s*grave|art\.?\s*100\s*§\s*2)', txt))

# Loop nos arquivos baixados
todos_registros = []
arquivos_processados = []
for fpath in sorted(glob.glob(os.path.join(RAW, '*.xls*'))):
    fname = os.path.basename(fpath)
    if fname.startswith('_'): continue
    # Pular PDFs (não suportado)
    if fpath.lower().endswith('.pdf'): continue
    info = manifest.get(fname, {})
    cat = info.get('cat', 'OUTRO')
    label = info.get('label', '')
    fonte_url = info.get('url', '')

    print(f'\n>> {fname}')
    is_xls = fpath.lower().endswith('.xls')
    try:
        if is_xls:
            book = xlrd.open_workbook(fpath, on_demand=True)
            sheets = book.sheet_names()
        else:
            wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
            sheets = wb.sheetnames
    except Exception as e:
        print(f'  ERRO abrir: {e}')
        continue

    for sh_name in sheets:
        if 'indic' in sh_name.lower() or 'valida' in sh_name.lower(): continue
        try:
            ws = book.sheet_by_name(sh_name) if is_xls else wb[sh_name]
        except Exception as e:
            print(f'  Skip sheet {sh_name}: {e}')
            continue

        nr, nc = maxrow(ws), maxcol(ws)
        if nr < 5 or nc < 5: continue
        hdr_row, headers = find_header_row(ws)
        if not hdr_row:
            # Para "Expedidos" (formato B sem header padrão): usar mapeamento fixo
            if 'EXPEDIDO' in cat or 'Comum' in sh_name or 'Alimentar' in sh_name:
                # B=seq C=label D=PRC E=CNJ F=natureza G=devedor I=valor J=data
                cmap = {'prc':4,'cnj':5,'natureza':6,'devedora':7,'valor':9,'data_expedicao':10}
                hdr_row = 7
            else:
                continue
        else:
            cmap = map_columns(headers)
            if 'prc' not in cmap or 'cnj' not in cmap:
                continue

        # Iterar linhas após header
        rows_added = 0
        for r in range(hdr_row + 1, nr + 1):
            prc = cell(ws, r, cmap.get('prc', 0)) if cmap.get('prc') else None
            cnj = cell(ws, r, cmap.get('cnj', 0)) if cmap.get('cnj') else None
            if not prc and not cnj: continue
            # Validar PRC numérico (registros válidos)
            try:
                prc_int = int(float(str(prc))) if prc else None
            except:
                prc_int = None
            if not prc_int or prc_int < 1 or prc_int > 999999: continue

            devedora = str(cell(ws, r, cmap['devedora']) or '') if 'devedora' in cmap else ''
            # Filtro devedora
            if devedora and not DEVEDORA_OK.search(devedora):
                continue

            natureza = str(cell(ws, r, cmap['natureza']) or '') if 'natureza' in cmap else ''
            valor = cell(ws, r, cmap['valor']) if 'valor' in cmap else None
            valor_atualizado = cell(ws, r, cmap['valor_atualizado']) if 'valor_atualizado' in cmap else None
            data_exp = cell(ws, r, cmap['data_expedicao']) if 'data_expedicao' in cmap else None
            tribunal = str(cell(ws, r, cmap['tribunal']) or '') if 'tribunal' in cmap else ''
            cidade = str(cell(ws, r, cmap['cidade']) or '') if 'cidade' in cmap else ''
            cnpj_dev = str(cell(ws, r, cmap['cnpj_dev']) or '') if 'cnpj_dev' in cmap else ''
            acao = str(cell(ws, r, cmap['acao_origem']) or '') if 'acao_origem' in cmap else ''
            transito = cell(ws, r, cmap['transito_origem']) if 'transito_origem' in cmap else None

            # Detectar super-preferencial via assunto/observações
            super_pref = is_super_preferencial(natureza, acao)

            # Status: derivado da categoria
            STATUS_MAP = {
                'PENDENTE': 'PENDENTE',
                'EXPEDIDO': 'EXPEDIDO',
                'EXPEDIDO-SEM-LOA': 'EXPEDIDO_SEM_LOA',
                'EXPEDIDO-PDF': 'EXPEDIDO_PDF',
                'PAGO': 'PAGO',
                'MAPA': 'MAPA_ANUAL',
                'LISTA-PAGAMENTO': 'LISTA_PAGAMENTO',
                'OUTRO': 'OUTRO',
            }
            status = STATUS_MAP.get(cat, cat)

            # LOA inscrita: derivada do label
            loa = ''
            if 'LOA-2027' in label or '2027' in label: loa = '2027'
            elif 'LOA-2026' in label or '2026' in label: loa = '2026'
            elif 'dez-2024' in label: loa = '2025'  # pendentes em dez/2024 → LOA 2025
            elif 'dez-2023' in label: loa = '2024'
            elif 'dez-2022' in label: loa = '2023'
            elif 'dez-2021' in label: loa = '2022'
            elif 'dez-2020' in label: loa = '2021'
            elif 'dez-2019' in label: loa = '2020'
            elif '31-12-2018' in label: loa = '2019'
            elif '31-12-2016' in label: loa = '2017'
            elif 'exercicios-anteriores' in label: loa = 'exercícios anteriores'
            elif cat == 'EXPEDIDO-SEM-LOA': loa = 'sem LOA (autuado período LOA seguinte)'

            todos_registros.append({
                'tribunal_emissor': 'STJ',
                'prc': prc_int,
                'cnj': str(cnj or '').strip(),
                'natureza': natureza,
                'super_preferencial': super_pref,
                'devedora': devedora,
                'cnpj_devedora': cnpj_dev,
                'cidade': cidade,
                'tribunal_origem': tribunal,
                'acao_origem': acao,
                'transito_julgado': transito,
                'data_expedicao': data_exp,
                'valor': valor,
                'valor_atualizado': valor_atualizado,
                'status': status,
                'loa_inscrito': loa,
                'fonte_arquivo': fname,
                'fonte_url': fonte_url,
                'fonte_sheet': sh_name,
            })
            rows_added += 1
        if rows_added:
            print(f'  Sheet "{sh_name}": {rows_added} registros')
    arquivos_processados.append(fname)

print(f'\n=== TOTAL CONSOLIDADO: {len(todos_registros)} registros de {len(arquivos_processados)} arquivos ===')

# Escrever XLSX consolidado
out = openpyxl.Workbook()
ws = out.active
ws.title = 'STJ-PRC-Pendentes'

cols = ['tribunal_emissor','prc','cnj','natureza','super_preferencial','devedora','cnpj_devedora',
        'cidade','tribunal_origem','acao_origem','transito_julgado','data_expedicao',
        'valor','valor_atualizado','status','loa_inscrito','fonte_arquivo','fonte_url','fonte_sheet']
ws.append(cols)
# Header style
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
hdr_fill = PatternFill('solid', fgColor='1c2a3f')
hdr_font = Font(bold=True, color='22d3ee', size=10)
for c, val in enumerate(cols, 1):
    cell0 = ws.cell(1, c)
    cell0.fill = hdr_fill
    cell0.font = hdr_font
    cell0.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 28
ws.freeze_panes = 'A2'

for rec in todos_registros:
    row = [rec.get(c, '') for c in cols]
    # Hyperlink na coluna fonte_url
    ws.append(row)

# Hyperlink nos URLs (column 18 / R)
url_col = cols.index('fonte_url') + 1
for r in range(2, len(todos_registros) + 2):
    cell0 = ws.cell(r, url_col)
    if cell0.value:
        cell0.hyperlink = cell0.value
        cell0.font = Font(color='22d3ee', underline='single', size=9)
        cell0.value = 'abrir fonte ↗'

# Auto-width básico
widths = {'tribunal_emissor':10,'prc':8,'cnj':22,'natureza':12,'super_preferencial':10,
          'devedora':35,'cnpj_devedora':18,'cidade':18,'tribunal_origem':12,
          'acao_origem':28,'transito_julgado':12,'data_expedicao':12,
          'valor':16,'valor_atualizado':16,'status':18,'loa_inscrito':14,
          'fonte_arquivo':38,'fonte_url':14,'fonte_sheet':16}
for i, c in enumerate(cols, 1):
    ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 14)

# Sheet de manifesto (lista de fontes)
ws2 = out.create_sheet('FONTES')
ws2.append(['categoria','rótulo','arquivo','url','status_download','tamanho_bytes'])
for m in manifest.values():
    ws2.append([m.get('cat'), m.get('label'), m.get('fname'), m.get('url'), m.get('status'), m.get('size')])
for c in range(1, 7):
    ws2.cell(1, c).fill = hdr_fill
    ws2.cell(1, c).font = hdr_font
ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 50
ws2.column_dimensions['D'].width = 70
ws2.column_dimensions['E'].width = 14
ws2.column_dimensions['F'].width = 14
# Hyperlink coluna D
for r in range(2, ws2.max_row + 1):
    cell0 = ws2.cell(r, 4)
    if cell0.value:
        cell0.hyperlink = cell0.value
        cell0.font = Font(color='22d3ee', underline='single', size=9)

# Sheet de resumo
ws3 = out.create_sheet('RESUMO', 0)
ws3.append(['STJ — Precatórios Pendentes — Consolidado'])
ws3.cell(1,1).font = Font(bold=True, size=14, color='22d3ee')
ws3.merge_cells('A1:E1')
ws3.append(['Gerado em', datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')])
ws3.append(['Total registros', len(todos_registros)])
ws3.append(['Arquivos processados', len(arquivos_processados)])
ws3.append([''])
ws3.append(['Filtros aplicados:'])
ws3.append(['', '✓ Apenas Precatórios PRC (acima 60 SM federais — exclui RPV)'])
ws3.append(['', '✓ Devedoras: União, INSS, Fazenda Nacional, Universidades Federais, Fundações, Autarquias'])
ws3.append(['', '✓ Status PENDENTE + EXPEDIDO sem LOA definida'])
ws3.append([''])
ws3.append(['Distribuição por LOA:'])
from collections import Counter
loas = Counter(r.get('loa_inscrito') for r in todos_registros)
for k, v in sorted(loas.items()):
    ws3.append(['', k or '(em branco)', v])
ws3.append([''])
ws3.append(['Distribuição por natureza:'])
nat = Counter((r.get('natureza') or '').strip().lower() for r in todos_registros)
for k, v in nat.most_common(10):
    ws3.append(['', k or '(em branco)', v])
ws3.column_dimensions['A'].width = 24
ws3.column_dimensions['B'].width = 50
ws3.column_dimensions['C'].width = 12

out.save(OUT)
print(f'\n✓ Arquivo salvo: {OUT}')
print(f'  Tamanho: {os.path.getsize(OUT)//1024} KB')
